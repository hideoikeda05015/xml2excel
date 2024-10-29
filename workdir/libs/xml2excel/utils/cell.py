import importlib
import openpyxl # type: ignore
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers # type: ignore
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU, pixels_to_points, EMU_to_pixels, dxa_to_inch # type: ignore
from openpyxl.drawing.xdr import XDRPositiveSize2D # type: ignore
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, OneCellAnchor, AnchorMarker # type: ignore
from openpyxl.styles import Font, Color, colors # type: ignore
from libs.xml2excel.utils import cell, util
for _m in [util, cell]:
    importlib.reload(_m)

def setBoxBorder(self, ws, row_start_pos, row_len, col_start_pos, col_len, border_color="", number="", number_color=""):

    for _row in range(row_start_pos, row_start_pos + row_len):
        for _col in range(col_start_pos, col_start_pos + col_len):
            target_cell = ws.cell(row=_row, column=_col)

            if border_color == "":
                border_color = self._border_color
            
            if _row == row_start_pos:
                _top_style = 'thin'
                _top_color = border_color
            elif target_cell.border.top.color is None:
                _top_style = None
                _top_color = None
            else :
                _top_style = target_cell.border.top.style
                _top_color = target_cell.border.top.color
            
            if _row == row_start_pos + row_len - 1:
                _bottom_style = 'thin'
                _bottom_color = border_color
            elif target_cell.border.bottom.color is None: # なぜかstyleだとbottomだけthinがある
                _bottom_style = None
                _bottom_color = None
            else :
                _bottom_style = target_cell.border.bottom.style
                _bottom_color = target_cell.border.bottom.color
            
            if _col == col_start_pos:
                _left_style = 'thin'
                _left_color = border_color
            elif target_cell.border.left.color is None:
                _left_style = None
                _left_color = None
            else :
                _left_style = target_cell.border.left.style
                _left_color = target_cell.border.left.color
            
            if _col == col_start_pos + col_len - 1:
                _right_style = 'thin'
                _right_color = border_color
            elif target_cell.border.right.color is None:
                _right_style = None
                _right_color = None
            else :
                _right_style = target_cell.border.right.style
                _right_color = target_cell.border.right.color

            side_top = Side(style=_top_style, color=_top_color)
            side_bottom = Side(style=_bottom_style, color=_bottom_color)
            side_left = Side(style=_left_style, color=_left_color)
            side_right = Side(style=_right_style, color=_right_color)

            target_cell.border = Border(top=side_top, bottom=side_bottom, left=side_left, right=side_right)

    if number != "":
        _tmp_image_path = util.createText2ImagePNG(self, number, number_color)
        image = openpyxl.drawing.image.Image(_tmp_image_path)
        col_offset = -1 * pixels_to_EMU(self._width_point_size / 2 + 2)
        row_offset = -1 * pixels_to_EMU(self._height_point_size / 2 - 1)
        size_ext = XDRPositiveSize2D(pixels_to_EMU(12), pixels_to_EMU(12))
        maker = AnchorMarker(col=col_start_pos + col_len - 1, colOff=col_offset, row=row_start_pos - 1, rowOff=row_offset)
        image.anchor = OneCellAnchor(_from=maker, ext=size_ext)
        ws.add_image(image) 
        self.diagram_number = self.diagram_number + 1

def setCellwithBorder(self, ws, row_start_pos, row_len, col_start_pos, col_len, color, text_aling, value, is_noborder=False, link=""):
    target_cell = ws.cell(row=row_start_pos, column=col_start_pos)
    fill = PatternFill(patternType='solid', fgColor=color)
    target_cell.fill = fill

    _font_size = self._font_size
    if "font_size" in value:
        _font_size = int(value["font_size"])

    target_cell.font = Font(size=_font_size)

    if is_noborder != True:
        side1 = Side(style='thin', color=self._border_color)
        target_cell.border = Border(top=side1, bottom=side1, left=side1, right=side1)

    ws.merge_cells(start_row=row_start_pos, start_column=col_start_pos, end_row=row_start_pos + row_len - 1, end_column=col_start_pos + col_len - 1)

    if link != "":
        #if link[0] != "#" and "!" not in link:
        #    link = "#" + link + "!A1"
        target_cell.hyperlink = link
        #target_cell.style = "Hyperlink"
        target_cell.font = Font(u='single', color=colors.BLUE, size=_font_size)
    if value["value"] != "" and value["value"].isdigit():
        target_cell.number_format = numbers.FORMAT_NUMBER
        target_cell.value = int(value["value"])
    else :
        target_cell.value = value["value"]
    target_cell.alignment = Alignment(horizontal=text_aling, vertical='center', wrapText=True)
    #if link != "":
    #    target_cell.value = '=HYPERLINK("' + link + '","' + value["value"] + '")'

def setCellwithBorder_Numnber(self, sheet, row_start_pos, row_len, col_start_pos, col_len, color, text_aling, value, is_noborder=True, link="", number="", number_color=""):
    setCellwithBorder(self, sheet, row_start_pos, row_len, col_start_pos, col_len, color, text_aling, value, is_noborder, link)

    if number != "":
        _tmp_image_path = util.createText2ImagePNG(self, number, number_color)
        image = openpyxl.drawing.image.Image(_tmp_image_path)
        col_offset = -1 * pixels_to_EMU(self._width_point_size / 2 + 2)
        row_offset = -1 * pixels_to_EMU(self._height_point_size / 2 - 1)
        size_ext = XDRPositiveSize2D(pixels_to_EMU(12), pixels_to_EMU(12))
        maker = AnchorMarker(col=col_start_pos + col_len - 1, colOff=col_offset, row=row_start_pos - 1, rowOff=row_offset)
        image.anchor = OneCellAnchor(_from=maker, ext=size_ext)
        sheet.add_image(image) 
        self.diagram_number = self.diagram_number + 1

def setCell(self, ws, row_start_pos, row_len, col_start_pos, col_len, color, text_aling, value, is_noborder=True, link=""):
    setCellwithBorder(self, ws, row_start_pos, row_len, col_start_pos, col_len, color, text_aling, value, is_noborder, link)

def setInlineTitle(self, sheet, start_pos, title, default_start_row=-1):

    if default_start_row == -1:
        default_start_row = self._start_col + self._body_lr_padding + self._page_lr_padding
    else :
        default_start_row = default_start_row + self._body_lr_padding + self._page_lr_padding

    pos = start_pos
    if title["value"] != "":
        target_cell = sheet.cell(row=start_pos, column=default_start_row)
        fill = PatternFill(patternType='solid', fgColor=self._background_color)
        target_cell.fill = fill
        target_cell.font = Font(size=self._font_size)

        _tmp_len = len(title["value"]) + 2
        if _tmp_len == 0:
            _tmp_len = self._inline_title_length
        elif _tmp_len > self._width_size - self._body_lr_padding * 2 - self._page_lr_padding * 2 - default_start_row:
            _tmp_len = self._width_size - self._body_lr_padding * 2 - self._page_lr_padding * 2 - default_start_row

        sheet.merge_cells(start_row=start_pos, start_column=default_start_row, end_row=start_pos, end_column=default_start_row + _tmp_len)

        side1 = Side(style='thin', color=self._border_color)
        for col in range(default_start_row, default_start_row + self._width_size - self._body_lr_padding * 2 - self._page_lr_padding * 2):
            _cell = sheet.cell(row=start_pos, column=col)
            _cell.border = Border(bottom=side1)

        if title["value"] != "" and title["value"].isdigit():
            target_cell.number_format = numbers.FORMAT_NUMBER
            target_cell.value = int(title["value"])
        else :
            target_cell.value = title["value"]
        target_cell.alignment = Alignment(horizontal="left", vertical='center', wrapText=True)

        pos = start_pos + 2

    return pos